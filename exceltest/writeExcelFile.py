import openpyxl
def outputExcel(inputDetail,outputExcelPath):
  #Xác định số hàng và cột lớn nhất trong file excel cần tạo
    row = len(inputDetail)
    column = len(inputDetail[0])
  #Tạo một workbook mới và active nó
    wb = openpyxl.Workbook()
    ws = wb.active
  
  #Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
    for i in range(0,row):
    # for j in range(0,column):
    #   v = inputDetail[i][j]
    #   ws.cell(column=j+1, row=i+1, value=v)
        v = inputDetail[i]
        ws.cell(column=1, row=i+1, value=v)

  #Lưu lại file Excel
    wb.save(outputExcelPath)