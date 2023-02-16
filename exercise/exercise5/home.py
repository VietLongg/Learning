from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtWidgets import QMessageBox
import sys
import FilterFile
import openpyxl

app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()

class FileHandler:
    listPaths = list()

    def __init__(self) -> None:
        pass

    def Filter(self, recListPaths, recFolderPaths, recFileExt, recCreatedDate):
        global retDict
        retDict = dict()
        for item in recListPaths:
            item = item[0]
            fileExt, createdDate = item[item.rfind('.') + 1:].split('|')
            if recFolderPaths == item[:item.find('_')] and recFileExt == fileExt and recCreatedDate == createdDate:
                folderPath, fileName = item[:item.rfind('|')].rsplit('/', 1)
                if fileName not in retDict:
                    retDict[fileName] = list()
                retDict[fileName].append(folderPath)
        return(retDict)

    def readListPath(self, recFilePath):
        wb = openpyxl.load_workbook(recFilePath)
        sheet = wb['Sheet']
        self.listPaths = list(sheet.values)
        
    def mainUi(self):
        global ui
        ui = FilterFile.Ui_MainWindow()
        ui.setupUi(MainWindow)
        ui.btn_Search.clicked.connect(self.search)
        ui.btn_ImportExcel.clicked.connect(self.importFromExcel)
        ui.btn_ExportExcel.clicked.connect(self.exportToExcel)
        MainWindow.show()

    def importFromExcel(self):
        InputExcelPath = ui.text_InputExcelPath.toPlainText()
        self.readListPath(InputExcelPath)
        self.showPopupImport()
    
    def exportToExcel(self):
        OutputExcelPath = ui.text_OutputExcelPath.toPlainText()
        self.outputExcel(retDict, OutputExcelPath)
        self.showPopupExport()

    def search(self):
        FolderPath = ui.text_InputFolder.toPlainText()
        InputExtension = ui.text_InputExtension.toPlainText()
        Date = ui.text_InputDate.toPlainText()
        self.Filter(self.listPaths, FolderPath, InputExtension, Date)
        self.showInTable()
    
    def outputExcel(self, inputDetail,outputExcelPath):
        row = len(inputDetail)

        wb = openpyxl.Workbook()
        ws = wb.active

        j = 0
        for i in inputDetail:
            v = str(i) + ' | ' + str(inputDetail[i])
            ws.cell(column = 1, row = j + 1, value = v)
            j += 1

        wb.save(outputExcelPath)

    def showInTable(self):
        row = len(retDict)
        j = 0
        ui.tableWidget.clear()
        for i in retDict:
            ui.tableWidget.setItem(j,0, QtWidgets.QTableWidgetItem(i))
            ui.tableWidget.setItem(j,1, QtWidgets.QTableWidgetItem(str(retDict[i])))
            j +=1

    def showPopupImport(self):
        msg = QMessageBox()
        msg.setWindowTitle('Message')
        msg.setText('Import Success')
        msg.setInformativeText('Press Yes please :))')
        msg.setDetailedText('Path: /home/long/Desktop/long/excel/listPaths.xlsx')
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Yes|QMessageBox.Cancel)
        msg.setDefaultButton(QMessageBox.Yes)
        x = msg.exec_()

    def showPopupExport(self):
        msg = QMessageBox()
        msg.setWindowTitle('Message')
        msg.setText('Export Success')
        msg.setIcon(QMessageBox.Information)
        x = msg.exec_()

if __name__ == "__main__":
    a = FileHandler()
    a.mainUi()
    sys.exit(app.exec_())