import openpyxl
class FileHandler:
    
    def __init__(self) -> None:
        pass
    
    def FilterFile(self, recListPaths, recFileExt):
        retDict = dict()
        for item in recListPaths:
            if recFileExt == item[item.rfind('.')+1:]:
                filePath, fileName = item.rsplit('/', 1)
                if fileName not in retDict:
                    retDict[fileName] = list()
                retDict[fileName].append(filePath)
        return(retDict)

    def Filter(self, recListPaths, recFolderPaths, recFileExt, recCreatedDate):
        retDict = dict()

        for item in recListPaths:
            item = item[0]
            fileExt, createdDate = item[item.rfind('.') + 1:].split('|')
            if recFolderPaths == item[:item.find('_')] and recFileExt == fileExt and recCreatedDate == createdDate:
                folderPath, fileName = item[:item.rfind('|')].rsplit('/', 1)
                if fileName not in retDict:
                    retDict[fileName] = list()
                retDict[fileName].append(folderPath)
        return(retDict)

    def readListPath(self, recFilePath):
        wb = openpyxl.load_workbook(recFilePath)
        # print(wb.sheetnames)
        sheet = wb['Sheet']
        listPaths = list(sheet.values)
        return listPaths

    def outputExcel(self, inputDetail,outputExcelPath):
        row = len(inputDetail)
        # column = len(inputDetail[0])

        wb = openpyxl.Workbook()
        ws = wb.active

        j = 0
        for i in inputDetail:
            v = str(i) + ' | ' + str(inputDetail[i])
            ws.cell(column = 1, row = j + 1, value = v)
            j += 1

        wb.save(outputExcelPath)

if __name__ == "__main__":
    a = FileHandler()

    filePath = '/home/long/Desktop/long/excel/listPaths.xlsx'
    listPaths = a.readListPath(filePath)

    extSearch = input("Enter your file extension: ")
    ret = a.Filter(listPaths, 'C:/Folder', extSearch, '02/09/23')
    print(ret)
    outputExcelPath = '/home/long/Desktop/long/excel/Filter.xlsx'
    a.outputExcel(ret, outputExcelPath)
    print('Export ' + outputExcelPath[outputExcelPath.rfind('/') + 1:] + ' succeeded')